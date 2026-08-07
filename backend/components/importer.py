"""
Importer helpers para carga masiva de componentes desde CSV y GeoJSON.

Patrón:
  - `parse_csv(file_bytes)` y `parse_geojson(file_bytes)` leen el archivo y
    lo convierten en una lista de `ParsedComponent` dicts con TODOS los FK
    resueltos a IDs (district, type, operational_status, physical_status,
    criticality). Las coords vienen como lista de dicts:
        { 'easting': ..., 'northing': ..., 'srid_origin': int, 'criticality': int }
    o bien:
        { 'longitude': ..., 'latitude': ..., 'criticality': int }
    según venga en el archivo.

  - `ImportError` agrupa todos los errores de una fila/feature con su
    contexto (fila, código, mensaje). El caller decide cómo persistir.

  - Las FKs se resuelven por nombre (mayúsculas) o ubigeo, NUNCA por id.
    El usuario no conoce IDs; el archivo trae nombres legibles.

  - Para CSV de líneas (varios vértices por componente): el usuario
    repite filas con MISMO `code` y distinto `easting`/`northing`. El
    parser agrupa filas por `code` y produce un solo `ParsedComponent`
    con N coord entries. El `type` debe ser un tipo línea
    (`linea-conduccion`/`linea-aduccion`); si no, se reporta error.

Restricciones:
  - Encoding esperado: UTF-8 (acepta BOM vía utf-8-sig).
  - Si una fila/feature viola unique (district+type+code ya existe),
    se reporta error. El caller aborta la transacción completa y NO
    persiste NADA (rollback), reportando al usuario dónde está el
    conflicto. Es el patrón "abort-all-on-duplicate" elegido.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

import re
import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

from django.contrib.gis.geos import Point
from rest_framework.exceptions import ValidationError

from components.models import (
    Component,
    ComponentType,
    Criticality,
    OperationalStatus,
    PhysicalStatus,
)
from places.models import District
from core_shared.helpers import SpatialHelper

__all__ = [
    "ImportError",
    "ParsedComponent",
    "parse_csv",
    "parse_geojson",
    "parse_xlsx",
    "persist_components",
    "build_xlsx_template",
]


# ── Regex de validación de código (alineados con components/validators.py) ──
# `component_code_validator` exige `^\d{4}$` (4 dígitos numéricos).
# `operational_status_code_validator` exige `^\d{3}$` (3 dígitos).
# `physical_status_code_validator` exige `^[A-Z]{1}$` (1 letra mayúscula).
# Validamos en el importer ANTES de tocar la DB para que el usuario reciba
# un error claro por fila ("code 'CPT-001' debe ser 4 dígitos") en vez del
# genérico "value too long for type character varying(4)" que escupe
# PostgreSQL al violar max_length.
CODE_RE = re.compile(r'^\d{4}$')
OP_STATUS_RE = re.compile(r'^\d{3}$')
FIS_STATUS_RE = re.compile(r'^[A-Z]{1}$')

# Rangos UTM de Perú (alineados con SpatialHelper.utm_to_wgs84). Si una
# coordenada está fuera de rango, el parser lo reporta por fila en vez
# de esperar a que SpatialHelper reviente durante el persist.
UTM_EASTING_MIN = 100000.0
UTM_EASTING_MAX = 900000.0
UTM_NORTHING_MIN = 0.0
UTM_NORTHING_MAX = 10000000.0


# ── Tipos_linea definidos en el frontend como TipoComponente; en backend
# se mapean via el NAME del ComponentType. Aceptamos mayúsculas Según el
# esquema de `seed_*` puesto por el backend (`ComponentType.name`):
LINEA_TYPE_NAMES = {
    "LÍNEA DE CONDUCCIÓN",
    "LINEA DE CONDUCCION",
    "LÍNEA DE ADUCCIÓN",
    "LINEA DE ADUCCION",
}


@dataclass
class ImportError:
    """Error de una fila/feature con contexto para reportar al usuario."""
    row: int | None  # 1-based para CSV; idx para GeoJSON
    code: str | None
    message: str


@dataclass
class ParsedComponent:
    """Componente parseado listo para persistir (FKs resueltas a IDs)."""
    code: str
    name: str
    type_id: int
    district_ubigeo: str
    operational_status_code: str | None
    physical_status_code: str | None
    specification: str | None
    coords: list[dict[str, Any]] = field(default_factory=list)


def _validate_utm_range(
    easting: float,
    northing: float,
    row: int | None,
    vertice_idx: int | None = None,
) -> list[ImportError]:
    """
        Valida que un par UTM (easting, northing) caiga dentro del rango
        métrico válido para Perú ( SpatialHelper.utm_to_wgs84 ). Retorna
        lista con el ImportError si está fuera de rango, lista vacía si OK.
        `vertice_idx` (1-based) opcional para mensaje más claro en líneas.
    """
    errs: list[ImportError] = []
    v_label = f" vértice {vertice_idx}" if vertice_idx else ""
    if not (UTM_EASTING_MIN <= easting <= UTM_EASTING_MAX):
        errs.append(ImportError(
            row=row, code=None,
            message=(
                f"Fila{(' ' + str(row)) if row else ''}{v_label}: "
                f"easting='{easting}' fuera de rango UTM válido "
                f"[{int(UTM_EASTING_MIN)}, {int(UTM_EASTING_MAX)}]. "
                f"Verificá que el Este sea correcto (¿lo confundiste con el Norte?)."
            ),
        ))
    if not (UTM_NORTHING_MIN <= northing <= UTM_NORTHING_MAX):
        errs.append(ImportError(
            row=row, code=None,
            message=(
                f"Fila{(' ' + str(row)) if row else ''}{v_label}: "
                f"northing='{northing}' fuera de rango UTM válido "
                f"[{int(UTM_NORTHING_MIN)}, {int(UTM_NORTHING_MAX)}]."
            ),
        ))
    return errs


def _validate_component_fields(
    code: str,
    name: str,
    op_status_code: str | None,
    fis_status_code: str | None,
    row: int | None,
) -> list[ImportError]:
    """
        Validaciones de esquema ALINEADAS con `components/validators.py`:
          - `code`: debe matchear `^\\d{4}$` (4 dígitos numéricos).
          - `name`: no vacío.
          - `operational_status` (opcional): si presente, `^\\d{3}$`.
          - `physical_status` (opcional): si presente, `^[A-Z]{1}$`.

        Retorna lista de ImportError (vacía si todo OK). Al validar acá
        evitamos el error genérico "value too long for type character
        varying(4)" de PostgreSQL y reportamos al usuario la fila exacta.
    """
    errs: list[ImportError] = []
    if not CODE_RE.match(code):
        errs.append(ImportError(
            row=row, code=code,
            message=(
                f"code='{code}' debe ser exactamente 4 dígitos numéricos "
                f"(ej: '0001'). Tu código tiene {len(code)} carácter(es)."
            ),
        ))
    if not name or not name.strip():
        errs.append(ImportError(
            row=row, code=code,
            message="campo 'name' vacío u obligatorio.",
        ))
    if op_status_code and not OP_STATUS_RE.match(op_status_code):
        errs.append(ImportError(
            row=row, code=code,
            message=(
                f"operational_status='{op_status_code}' debe ser 3 dígitos "
                f"numéricos (ej: '001')."
            ),
        ))
    if fis_status_code and not FIS_STATUS_RE.match(fis_status_code):
        errs.append(ImportError(
            row=row, code=code,
            message=(
                f"physical_status='{fis_status_code}' debe ser 1 letra "
                f"mayúscula A-Z (ej: 'A')."
            ),
        ))
    return errs


def _resolve_fk(model, label: str, **lookup) -> int | str | None:
    """Busca un registro por `lookup` y retorna su PK; None si no existe."""
    try:
        return model.objects.get(**lookup).pk
    except model.DoesNotExist:
        return None


def _resolve_type(name: str) -> int:
    """Resuelve nombre del tipo → ID. Lanza ValueError si no existe."""
    n = name.strip().upper()
    qs = ComponentType.objects.filter(name__iexact=name)
    if not qs.exists():
        raise ValueError(f"Tipo de componente '{name}' no encontrado.")
    return qs.first().pk


def _resolve_criticality(name: str) -> int:
    """Resuelve nombre de criticidad → ID."""
    qs = Criticality.objects.filter(name__iexact=name)
    if not qs.exists():
        raise ValueError(f"Criticidad '{name}' no encontrada.")
    return qs.first().pk


def _resolve_district(ubigeo: str) -> str:
    """Resuelve ubigeo → ubigeo (sólo valida que el distrito exista)."""
    if not District.objects.filter(ubigeo=ubigeo).exists():
        raise ValueError(f"Distrito con ubigeo '{ubigeo}' no encontrado.")
    return ubigeo


def _parse_float(value: str, field_name: str, row: int | None) -> float:
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        raise ValueError(
            f"Fila {row}: campo '{field_name}'='{value}' no es numérico."
        )


# ──────────────────────────────────────────────────────────────────────
# CSV
# ──────────────────────────────────────────────────────────────────────

CSV_REQUIRED_HEADERS = [
    "code",
    "name",
    "type",
    "district_ubigeo",
    "criticality",
    "easting",
    "northing",
]
CSV_OPTIONAL_HEADERS = [
    "operational_status",
    "physical_status",
    "specification",
]


def parse_csv(
    file_bytes: bytes,
) -> tuple[list[ParsedComponent], list[ImportError]]:
    """
    Parsea un archivo CSV y retorna (componentes, errores).

    Filas con mismo (code, district, type) se agrupan en un solo
    componente (clave unique_together del backend); el `type`
    asociado determina si es válido tener múltiples vértices (debe ser
    línea de conducción/aducción).
    """
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=",", skipinitialspace=True)

    if reader.fieldnames is None:
        return [], [ImportError(row=1, code=None, message="CSV vacío.")]
    headers_lower = {h.strip().lower(): h for h in reader.fieldnames}
    missing = [h for h in CSV_REQUIRED_HEADERS if h not in headers_lower]
    if missing:
        return [], [ImportError(
            row=1,
            code=None,
            message=f"Headers faltantes: {', '.join(missing)}. "
                    f"Headers esperados: {', '.join(CSV_REQUIRED_HEADERS)}.",
        )]

    # Convertir cada row del reader en (row_idx, norm_dict). Las filas con
    # alguno de los campos obligatorios vacíos generan ImportError acá
    # mismo para no propagarlos al builder.
    rows: list[tuple[int, dict]] = []
    errors_pre: list[ImportError] = []
    for idx, row in enumerate(reader, start=2):
        norm = {k.strip().lower(): (str(v or "")).strip() for k, v in row.items() if k}
        rows.append((idx, norm))

    return _build_components_from_rows(rows, errors_pre)


# ──────────────────────────────────────────────────────────────────────
# Builder compartido (lo usan CSV y XLSX)
# ──────────────────────────────────────────────────────────────────────

def _build_components_from_rows(
    rows: list[tuple[int, dict]],
    pre_errors: list[ImportError] | None = None,
) -> tuple[list[ParsedComponent], list[ImportError]]:
    """
        Construye `ParsedComponent` a partir de rows ya normalizados
        (idx 1-based, dict lowercase keys).

        Estrategia:
          1. Saltear filas con `code` vacío o `code`='ELIMINAR' (marcador
             de plantilla XLSX).
          2. Agrupar por (code, district_ubigeo, type) — la clave natural
             unique_together del modelo Component. Vértices del mismo
             componente (línea) comparten esos 3 valores; dos
             componentes distintos con mismo code pero distinto type
             son dos grupos separados.
          3. Por grupo: resolver FKs, validar consistencia entre filas,
             validar tipo línea-vs-puntual, validar UTM+rangos, validar
             code/name/estados con regex, y armar el ParsedComponent.
    """
    grouped: dict[tuple, list[tuple[int, dict]]] = {}
    order: list[tuple] = []
    errors: list[ImportError] = list(pre_errors or [])

    for idx, norm in rows:
        code = norm.get("code", "")
        if not code:
            errors.append(ImportError(
                row=idx, code=None, message="Campo 'code' vacío."
            ))
            continue
        # Saltear filas marcadas como "ELIMINAR" en plantillas XLSX.
        if code.upper().strip() == "ELIMINAR":
            continue
        group_key = (code, norm.get("district_ubigeo", ""), norm.get("type", ""))
        if group_key not in grouped:
            grouped[group_key] = []
            order.append(group_key)
        grouped[group_key].append((idx, norm))

    # Convertir cada grupo en un ParsedComponent
    componentes: list[ParsedComponent] = []
    for group_key in order:
        code, district_key, type_key = group_key
        rows_grupo = grouped[group_key]
        first_idx, first = rows_grupo[0]
        try:
            type_name = first["type"]
            type_id = _resolve_type(type_name)
            district_ubigeo = _resolve_district(first["district_ubigeo"])
            crit_id = _resolve_criticality(first["criticality"])
        except ValueError as e:
            errors.append(ImportError(row=first_idx, code=code, message=str(e)))
            continue

        # Verificar consistencia entre filas del mismo grupo
        inconsistent = None
        for ridx, r in rows_grupo[1:]:
            if r["type"].upper() != first["type"].upper():
                inconsistent = f"Fila {ridx}: 'type'='{r['type']}' difiere de la primera fila ('{first['type']}')."
                break
            if r["district_ubigeo"] != first["district_ubigeo"]:
                inconsistent = f"Fila {ridx}: 'district_ubigeo'='{r['district_ubigeo']}' difiere de la primera fila ('{first['district_ubigeo']}')."
                break
        if inconsistent:
            errors.append(ImportError(row=first_idx, code=code, message=inconsistent))
            continue

        # Si hay varias filas, el tipo debe ser línea
        es_linea_prep = first["type"].upper() in LINEA_TYPE_NAMES
        if len(rows_grupo) > 1 and not es_linea_prep:
            errors.append(ImportError(
                row=first_idx,
                code=code,
                message=(
                    f"Tipo '{first['type']}' no admite múltiples vértices "
                    f"({len(rows_grupo)} filas). Sólo líneas de conducción/aducción."
                ),
            ))
            continue

        # Build coords
        coords: list[dict[str, Any]] = []
        for ridx, r in rows_grupo:
            try:
                easting = _parse_float(r["easting"], "easting", ridx)
                northing = _parse_float(r["northing"], "northing", ridx)
            except ValueError as e:
                errors.append(ImportError(row=ridx, code=code, message=str(e)))
                coords = []
                break
            range_errors = _validate_utm_range(
                easting, northing, row=ridx,
                vertice_idx=(len(coords) + 1) if len(rows_grupo) > 1 else None,
            )
            if range_errors:
                errors.extend(range_errors)
                coords = []
                break
            try:
                v_crit_id = _resolve_criticality(r["criticality"])
            except ValueError as e:
                errors.append(ImportError(row=ridx, code=code, message=str(e)))
                coords = []
                break
            coords.append({
                "criticality": v_crit_id,
                "easting": easting,
                "northing": northing,
                "srid_origin": 18,
            })

        if not coords:
            continue  # ya está en errors

        op_status_code = first.get("operational_status") or None
        fis_status_code = first.get("physical_status") or None
        spec = first.get("specification") or None

        val_errors = _validate_component_fields(
            code=code,
            name=first.get("name", ""),
            op_status_code=op_status_code,
            fis_status_code=fis_status_code,
            row=first_idx,
        )
        if val_errors:
            errors.extend(val_errors)
            continue

        componentes.append(ParsedComponent(
            code=code,
            name=first["name"],
            type_id=type_id,
            district_ubigeo=district_ubigeo,
            operational_status_code=op_status_code,
            physical_status_code=fis_status_code,
            specification=spec,
            coords=coords,
        ))

    return componentes, errors


# ──────────────────────────────────────────────────────────────────────
# XLSX (Excel)
# ──────────────────────────────────────────────────────────────────────

def parse_xlsx(
    file_bytes: bytes,
) -> tuple[list[ParsedComponent], list[ImportError]]:
    """
    Parsea un archivo .xlsx y retorna (componentes, errores).

    La plantilla (ver `build_xlsx_template`) tiene:
      - Fila 1: instrucciones generales.
      - Filas 2-N: descripción por columna (campo, requerido/opcional,
        formato, valores válidos). Estas filas son ignoradas por el parser
        — sólo buscamos la fila de headers.
      - Fila k: headers reales (code, name, type, district_ubigeo,
        criticality, easting, northing, operational_status,
        physical_status, specification).
      - Filas k+1 ...: datos. Filas con `code = ELIMINAR` se saltean
        (instrucción para el usuario de borrar las de ejemplo antes
        de subir).

    Cualquier fila de ejemplo dejada por el usuario con code=ELIMINAR
    se ignora automáticamente — el modal/base le avisa igual que debe
    eliminar esas filas antes de subir, pero el parser es tolerante.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return [], [ImportError(
            row=None, code=None,
            message=f"No se pudo leer el archivo XLSX: {e}",
        )]

    ws = wb.active
    if ws is None:
        return [], [ImportError(row=None, code=None, message="Archivo XLSX sin hojas.")]

    # Localizar la fila de headers exacta: primera fila cuyos valores
    # (lower-cased) contienen TODOS los headers requeridos. Esto
    # permite que la plantilla tenga filas de descripción arriba sin
    # romper el parseo.
    required_lower = set(CSV_REQUIRED_HEADERS)
    header_row_idx = None
    headers_map: dict[str, int] = {}  # header_lower → col_index (0-based)

    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        # row: tuple de N valores, algunos pueden ser None.
        row_lower = [
            (str(v).strip().lower() if v is not None else "")
            for v in row
        ]
        if required_lower.issubset({v for v in row_lower if v}):
            header_row_idx = ri
            for ci, val in enumerate(row_lower):
                if val in required_lower or val in CSV_OPTIONAL:
                    headers_map[val] = ci
            break

    if header_row_idx is None:
        return [], [ImportError(
            row=None, code=None,
            message=(
                "No se encontró fila de headers. La plantilla debe contener "
                f"una fila con los encabezados: {', '.join(CSV_REQUIRED_HEADERS)}."
            ),
        )]

    # Validar headers faltantes
    missing = [h for h in CSV_REQUIRED_HEADERS if h not in headers_map]
    if missing:
        return [], [ImportError(
            row=header_row_idx + 1,
            code=None,
            message=f"Headers faltantes: {', '.join(missing)}.",
        )]

    # Extraer rows: empezar después de header_row_idx. Cada row se
    # normaliza a un dict {header_lower: str_value} ignorando columnas
    # que no están en headers_map. Filas completamente vacías se saltean.
    rows: list[tuple[int, dict]] = []
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= header_row_idx + 1:
            continue
        # Saltar filas completamente vacías
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        norm: dict[str, str] = {}
        for header_lower, ci in headers_map.items():
            val = row[ci] if ci < len(row) else None
            if val is None:
                norm[header_lower] = ""
            elif isinstance(val, float) or isinstance(val, int):
                #openpyxl puede traer números como float; normalizamos
                # sin perder info (ej: 463529.0 → "463529").
                if isinstance(val, float) and val.is_integer():
                    norm[header_lower] = str(int(val))
                else:
                    norm[header_lower] = str(val)
            else:
                norm[header_lower] = str(val).strip()
        rows.append((ri, norm))

    wb.close()
    return _build_components_from_rows(rows)


# ──────────────────────────────────────────────────────────────────────
# Plantilla Excel
# ──────────────────────────────────────────────────────────────────────

def build_xlsx_template() -> bytes:
    """
        Genera el XLSX plantilla con:
          - Hoja 'Cargar Componentes': instrucciones + descripción por
            columna + fila de headers + filas de ejemplo (marcadas con
            code='ELIMINAR').
          - Hoja 'Valores válidos': listados de tipos, criticidades y
            estados operacionales/físicos del backend para que el usuario
            sepa qué valores acepta cada columna.

        El usuario debe:
          1. Eliminar las filas de ejemplo (code='ELIMINAR').
          2. Llenar sus datos en la grilla manteniendo los headers.
          3. Subir el archivo al backend.

        `parse_xlsx` identifica la fila de headers automáticamente y
        saltea filas con code='ELIMINAR', así que aunque el usuario
        olvide borrar las de ejemplo, el parser las ignora.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Cargar Componentes"

    # Estilos
    navy_fill = PatternFill("solid", fgColor="070B5B")
    white_bold = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    bold = Font(bold=True, name="Calibri")
    small_italic = Font(italic=True, color="6F6C8F", size=9, name="Calibri")
    normal = Font(name="Calibri", size=10)
    thin = Side(border_style="thin", color="ABB5BE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Fila 1: instrucciones ──
    ws["A1"] = (
        "INSTRUCCIONES: Complete UNA fila por componente (o por vértice si es línea). "
        "Elimine las filas de ejemplo (code=ELIMINAR) antes de subir. "
        "Consulse la hoja 'Valores válidos' para conocer los valores permitidos."
    )
    ws["A1"].font = small_italic
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.row_dimensions[1].height = 30

    # ── Filas 2-3: descripción por columna ──
    # Encabezado de "descriptor" en columna K, "valor" en L (fuera de la
    # grilla de datos). Pero más claro: dejamos la descripción en filas
    # puestas arriba de cada header.
    # Estructura más simple:
    #   Fila 2: descripción de cada columna (qué formato/values).
    #   Fila 3: headers reales (code, name, ...).
    #   Fila 4+: ejemplos con code=ELIMINAR.

    descriptions = [
        ("code",
         "4 dígitos numéricos. Ej: '0001'.\n"
         "Requerido. Unique por (district + type + code)."),
        ("name",
         "Texto hasta 50 caracteres.\nRequerido. Ej: 'CAPTACION PUCUSANI'."),
        ("type",
         "Nombre del tipo de componente (ver hoja 'Valores válidos').\n"
         "Requerido. Ej: 'CAPTACIÓN', 'RESERVORIO', 'LÍNEA DE CONDUCCIÓN'."),
        ("district_ubigeo",
         "6 dígitos ubigeo del distrito. Requerido. Ej: '120303'.\n"
         "Valores válidos: branches activas del backend."),
        ("criticality",
         "ALTA / MEDIA / BAJA. Requerido. Por vértice."),
        ("easting",
         "Número UTM Este (metros). Rango válido: 100000–900000.\n"
         "Requerido. Si es línea, una fila por vértice."),
        ("northing",
         "Número UTM Norte (metros). Rango válido: 0–10000000.\n"
         "Requerido."),
        ("operational_status",
         "3 dígitos, código del estado operativo. Opcional.\n"
         "Valores: ver hoja 'Valores válidos'."),
        ("physical_status",
         "1 letra mayúscula A-Z. Opcional.\n"
         "Valores: ver hoja 'Valores válidos'."),
        ("specification",
         "Texto libre (max 300). Opcional; puede quedar vacío (= NULL)."),
    ]

    # Descripciones (fila 2)
    for ci, (_, desc) in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=ci, value=desc)
        cell.font = small_italic
        cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )
        cell.border = border
    ws.row_dimensions[2].height = 70

    # Headers reales (fila 3)
    headers = [h for h, _ in descriptions]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 24

    # ── Filas de ejemplo (marcadas con code='ELIMINAR') ──
    examples = [
        ("ELIMINAR", "Captación Río Pichanaqui (EJEMPLO)", "CAPTACIÓN",
         "120303", "ALTA", 506961, 8788264, "001", "B",
         "Captación superficial — borre esta fila antes de subir"),
        ("ELIMINAR", "Reservorio San Ramón (EJEMPLO)", "RESERVORIO",
         "120305", "MEDIA", 465120, 8779850, "001", "B", ""),
        ("ELIMINAR", "Estación Bombeo Satipo (EJEMPLO)",
         "ESTACIÓN DE BOMBEO Y REBOMBEO DE AGUA POTABLE",
         "120601", "ALTA", 471200, 8780050, "002", "C", ""),
        # Línea con 2 vértices: repetir code en 2 filas.
        ("ELIMINAR", "Línea Conducción Tramo 1 (EJEMPLO)", "LÍNEA DE CONDUCCIÓN",
         "120303", "ALTA", 463600, 8777300, "001", "A", "Tramo de ejemplo"),
        ("ELIMINAR", "Línea Conducción Tramo 1 (EJEMPLO)", "LÍNEA DE CONDUCCIÓN",
         "120303", "MEDIA", 463700, 8777350, "001", "A", "Tramo de ejemplo"),
    ]
    for ri, ex in enumerate(examples, start=4):
        for ci, v in enumerate(ex, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = normal
            cell.border = border

    # Anchos de columna
    widths = [12, 38, 42, 18, 12, 14, 14, 22, 18, 38]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data validations (dropdowns) ──
    # Necesitamos una 2da hoja con los listados vivos del backend; los
    # DataValidations tipo "list" los referencian por rango.
    #
    # IMPORTANTE: openpyxl no acepta `=` inicial en `formula1` para DVs;
    # la forma correcta es `'SheetName'!$A$2:$A$N` (comillas simples
    # porque openpyxl las necesita macOS-style y para nombres con
    # espacios/acentos). SIEMPRE deben tener comillas para evitar
    # errores con Excel también.
    #
    # Nombre de hoja ASCII "Valores" (sin espacios ni acentos). Eso evita
    # problemas de parsing de la fórmula en distintos lectores.
    ws_vals = wb.create_sheet("Valores")

    type_names = list(ComponentType.objects.values_list("name", flat=True))
    crit_names = list(Criticality.objects.values_list("name", flat=True))
    op_codes = list(OperationalStatus.objects.values_list("code", flat=True))
    fis_codes = list(PhysicalStatus.objects.values_list("code", flat=True))

    # Headers de la hoja "Valores"
    ws_vals["A1"] = "Tipos de componente"
    ws_vals["B1"] = "Criticidades"
    ws_vals["C1"] = "Estados operacionales (código = nombre)"
    ws_vals["D1"] = "Estados físicos (código = nombre)"
    for c in ["A", "B", "C", "D"]:
        ws_vals[f"{c}1"].font = white_bold
        ws_vals[f"{c}1"].fill = navy_fill
        ws_vals[f"{c}1"].alignment = Alignment(vertical="center")
    ws_vals.row_dimensions[1].height = 22

    # Pares (code, name) para mostrar leyenda legible.
    op_pairs = list(OperationalStatus.objects.values_list("code", "name"))
    fis_pairs = list(PhysicalStatus.objects.values_list("code", "name"))

    max_len = max(len(type_names), len(crit_names), len(op_codes), len(fis_codes), 1)
    for i in range(max_len):
        r = i + 2
        if i < len(type_names):
            ws_vals.cell(row=r, column=1, value=type_names[i])
        if i < len(crit_names):
            ws_vals.cell(row=r, column=2, value=crit_names[i])
        if i < len(op_codes):
            code = op_pairs[i][0] if i < len(op_pairs) else op_codes[i]
            name = op_pairs[i][1] if i < len(op_pairs) else ""
            ws_vals.cell(row=r, column=3, value=f"{code} = {name}")
        if i < len(fis_codes):
            code = fis_pairs[i][0] if i < len(fis_pairs) else fis_codes[i]
            name = fis_pairs[i][1] if i < len(fis_pairs) else ""
            ws_vals.cell(row=r, column=4, value=f"{code} = {name}")

    # Anchos de la hoja "Valores"
    ws_vals.column_dimensions["A"].width = 50
    ws_vals.column_dimensions["B"].width = 18
    ws_vals.column_dimensions["C"].width = 32
    ws_vals.column_dimensions["D"].width = 32

    # Helper para armar la formula de DV. Sin `=`, comillas simples por
    # consistencia. Sheet name "Valores" no las necesita pero da igual.
    def _dv_formula(col: str, count: int) -> str:
        return f"'Valores'!${col}$2:${col}${count + 1}"

    # Dropdowns en la hoja principal, filas 4 a 200 (margen para llenar).
    if type_names:
        # col B = type
        dv_type = DataValidation(
            type="list",
            formula1=_dv_formula("A", len(type_names)),
            allow_blank=False,
        )
        dv_type.add("B4:B200")
        ws.add_data_validation(dv_type)
    if crit_names:
        # col E = criticality
        dv_crit = DataValidation(
            type="list",
            formula1=_dv_formula("B", len(crit_names)),
            allow_blank=False,
        )
        dv_crit.add("E4:E200")
        ws.add_data_validation(dv_crit)
    if op_codes:
        # col H = operational_status
        dv_op = DataValidation(
            type="list",
            formula1=_dv_formula("C", len(op_codes)),
            allow_blank=True,
        )
        dv_op.add("H4:H200")
        ws.add_data_validation(dv_op)
    if fis_codes:
        # col I = physical_status
        dv_fis = DataValidation(
            type="list",
            formula1=_dv_formula("D", len(fis_codes)),
            allow_blank=True,
        )
        dv_fis.add("I4:I200")
        ws.add_data_validation(dv_fis)

    # Exportar a bytes (en memoria).
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────
# GeoJSON
# ──────────────────────────────────────────────────────────────────────

def parse_geojson(
    file_bytes: bytes,
) -> tuple[list[ParsedComponent], list[ImportError]]:
    """
    Parsea un archivo GeoJSON FeatureCollection.

    Cada Feature.Properties debe tener:
      - code (str)
      - name (str)
      - type (str)              → nombre del ComponentType (ej. "CAPTACIÓN")
      - district_ubigeo (str)  → ubigeo del distrito
      - criticality (str)      → nombre de criticidad (ej. "ALTA")
      - operational_status (str, opcional) → code del estado operativo
      - physical_status (str, opcional)    → code del estado físico
      - specification (str, opcional)

    Cada Feature.Geometry puede ser:
      - Point: 1 coordenada por componente.
      - LineString: N coordenadas (1 por vértice) → para líneas de
        conducción/aducción. Cada vértice se convierte en un ComponentCoord.
      - MultiPoint: igual que LineString (cada punto = un vértice).

    Coordenadas siempre en WGS84 [lng, lat] (estándar GeoJSON).
    Internamente convertimos a UTM para pasarlas al backend = igual que CSV.
    En realidad, el backend también acepta longitude/latitude directo
    en el ComponentCoordItemSerializer, así que pasamos lat/lng en vez
    de easting/northing para preservar precisión original.
    """
    try:
        data = json.loads(file_bytes.decode("utf-8-sig"))
    except json.JSONDecodeError as e:
        return [], [ImportError(row=None, code=None, message=f"GeoJSON inválido: {e}")]

    if data.get("type") != "FeatureCollection":
        return [], [ImportError(
            row=None, code=None,
            message="El archivo debe ser un GeoJSON FeatureCollection.",
        )]

    features = data.get("features", [])
    if not features:
        return [], [ImportError(row=None, code=None, message="FeatureCollection vacío.")]

    componentes: list[ParsedComponent] = []
    errors: list[ImportError] = []

    # Agrupar features por (code, district_ubigeo, type) — clave natural
    # unique_together del modelo Component. Igual que en CSV, esto
    # permite subir tipos mezclados con mismo `code` como componentes
    # distintos.
    grouped: dict[tuple, list[tuple[int, dict]]] = {}
    order: list[tuple] = []
    for idx, feat in enumerate(features, start=1):
        props = feat.get("properties", {}) or {}
        code = props.get("code")
        if not code:
            errors.append(ImportError(
                row=idx, code=None, message="Feature sin 'code' en properties."
            ))
            continue
        group_key = (code, props.get("district_ubigeo", ""), props.get("type", ""))
        if group_key not in grouped:
            grouped[group_key] = []
            order.append(group_key)
        grouped[group_key].append((idx, feat))

    for group_key in order:
        code, district_key, type_key = group_key
        feats = grouped[group_key]
        first_idx, first_feat = feats[0]
        props = first_feat.get("properties", {}) or {}

        try:
            type_name = props["type"]
            type_id = _resolve_type(type_name)
            district_ubigeo = _resolve_district(props["district_ubigeo"])
            crit_id = _resolve_criticality(props["criticality"])
        except KeyError as e:
            errors.append(ImportError(
                row=first_idx, code=code,
                message=f"Propiedad requerida faltante: {e}.",
            ))
            continue
        except ValueError as e:
            errors.append(ImportError(row=first_idx, code=code, message=str(e)))
            continue

        es_linea_prep = type_name.upper() in LINEA_TYPE_NAMES

        coords: list[dict[str, Any]] = []
        for fidx, feat in feats:
            props_local = feat.get("properties", {}) or {}
            geom = feat.get("geometry") or {}

            # ── Fuente de coordenadas ──
            # Preferimos UTM (East/Northing arrays en properties) porque es
            # la unidad de trabajo del editor. Si no están, fallback al
            # geometry.coordinates estándar GeoJSON (WGS84 [lng, lat]).
            utm_eastings = props_local.get("utm_eastings")
            utm_northings = props_local.get("utm_northings")
            utm_zone = props_local.get("utm_zone", 18)

            if utm_eastings is not None and utm_northings is not None:
                if not isinstance(utm_eastings, list) or not isinstance(utm_northings, list):
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message="'utm_eastings' y 'utm_northings' deben ser arrays numéricos.",
                    ))
                    coords = []
                    break
                if len(utm_eastings) != len(utm_northings):
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message="'utm_eastings' y 'utm_northings' tienen distinta longitud.",
                    ))
                    coords = []
                    break
                if len(utm_eastings) == 0:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message="Arrays 'utm_eastings'/'utm_northings' vacíos.",
                    ))
                    coords = []
                    break
                # Validar tipo línea vs 1 vértice
                if len(utm_eastings) > 1 and not es_linea_prep:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message=f"Tipo '{type_name}' no admite múltiples vértices; "
                                f"usa 1 solo en UTM arrays.",
                    ))
                    coords = []
                    break
                if len(utm_eastings) == 1 and es_linea_prep:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message="Tipo línea requiere al menos 2 vértices UTM.",
                    ))
                    coords = []
                    break
                for v_i, (e_str, n_str) in enumerate(zip(utm_eastings, utm_northings)):
                    try:
                        e = float(e_str)
                        n = float(n_str)
                    except (TypeError, ValueError):
                        errors.append(ImportError(
                            row=fidx, code=code,
                            message=f"Vértice {v_i + 1}: UTM no numérico (easting='{e_str}', northing='{n_str}').",
                        ))
                        coords = []
                        break
                    range_errors = _validate_utm_range(
                        e, n, row=fidx, vertice_idx=(v_i + 1),
                    )
                    if range_errors:
                        errors.extend(range_errors)
                        coords = []
                        break
                    coords.append({
                        "criticality": crit_id,
                        "easting": e,
                        "northing": n,
                        "srid_origin": int(utm_zone) if utm_zone else 18,
                    })
                if not coords:
                    break
                continue  # próximo feature del mismo code

            # Fallback: geometry.coordinates en WGS84 [lng, lat]
            gtype = geom.get("type")
            coordinates = geom.get("coordinates")
            if not coordinates:
                errors.append(ImportError(
                    row=fidx, code=code,
                    message="Feature sin 'geometry.coordinates' ni 'utm_eastings'/'utm_northings' en properties.",
                ))
                coords = []
                break

            # Normalizar a lista de [lng, lat]
            if gtype == "Point":
                pts: list[list[float]] = [coordinates]
                if es_linea_prep:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message="Tipo línea requiere LineString/MultiPoint o varios UTM, recibí Point.",
                    ))
                    coords = []
                    break
            elif gtype in ("LineString", "MultiPoint"):
                pts = coordinates
                if not es_linea_prep:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message=f"Tipo '{type_name}' no admite múltiples vértices; usa Point.",
                    ))
                    coords = []
                    break
            else:
                errors.append(ImportError(
                    row=fidx, code=code,
                    message=f"Geometry type '{gtype}' no soportada. Usa Point, LineString o MultiPoint, o bien 'utm_eastings'/'utm_northings' en properties.",
                ))
                coords = []
                break

            # Por cada vértice: lon/lat → ComponentCoord entry
            for v_i, pair in enumerate(pts):
                if len(pair) < 2:
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message=f"Vértice {v_i + 1} sin [lng, lat] válidos.",
                    ))
                    coords = []
                    break
                try:
                    lng = float(pair[0])
                    lat = float(pair[1])
                except (TypeError, ValueError):
                    errors.append(ImportError(
                        row=fidx, code=code,
                        message=f"Vértice {v_i + 1}: coordenadas no numéricas.",
                    ))
                    coords = []
                    break
                coords.append({
                    "criticality": crit_id,
                    "longitude": lng,
                    "latitude": lat,
                })
            if not coords:
                break

        if not coords:
            continue

        op_status_code = props.get("operational_status") or None
        fis_status_code = props.get("physical_status") or None
        spec = props.get("specification") or None

        # Validaciones de esquema (code/name/estados) — reporta errores
        # claros por feature en vez de que PostgreSQL reviente con
        # "value too long".
        val_errors = _validate_component_fields(
            code=code,
            name=props.get("name", ""),
            op_status_code=op_status_code,
            fis_status_code=fis_status_code,
            row=first_idx,
        )
        if val_errors:
            errors.extend(val_errors)
            continue

        componentes.append(ParsedComponent(
            code=code,
            name=props.get("name", code),  # fallback: code
            type_id=type_id,
            district_ubigeo=district_ubigeo,
            operational_status_code=op_status_code,
            physical_status_code=fis_status_code,
            specification=spec,
            coords=coords,
        ))

    return componentes, errors


# ──────────────────────────────────────────────────────────────────────
# Persistencia (dry_run=False) o preview (dry_run=True)
# ──────────────────────────────────────────────────────────────────────

def persist_components(
    componentes: list[ParsedComponent],
    dry_run: bool = False,
) -> tuple[int, list[ImportError]]:
    """
    Persiste N componentes transaccionalmente.

    Estrategia UNIQUE (district+type+code): el usuario eligió
    "abort-all-on-duplicate". Por eso:
      - Si ALGÚN componente parseado ya existe en la base (coincidencia
        exacta de district+type+code), abortamos la Tx completa y
        reportamos el conflicto. Nada se persiste.
      - Sino, hacemos bulk_create de Component + ComponentCoord.

    `dry_run=True` sólo valida (no persiste) — útil para preview.

    Retorna (creados, errores).
    """
    from components.models import ComponentCoord
    from django.db import transaction

    # 1. Validar duplicados intrínsecos (mismo code dentro del lote)
    seen = set()
    for c in componentes:
        key = (c.district_ubigeo, c.type_id, c.code)
        if key in seen:
            return 0, [ImportError(
                row=None, code=c.code,
                message=f"Duplicado interno: district={c.district_ubigeo} "
                        f"+ type_id={c.type_id} + code={c.code} aparece dos "
                        f"veces en el archivo.",
            )]
        seen.add(key)

    # 2. Validar duplicados contra la base existente
    existing_errors: list[ImportError] = []
    for c in componentes:
        exists = Component.objects.filter(
            district__ubigeo=c.district_ubigeo,
            type_id=c.type_id,
            code=c.code,
        ).exists()
        if exists:
            existing_errors.append(ImportError(
                row=None, code=c.code,
                message=f"Ya existe un componente con district="
                        f"{c.district_ubigeo}, type_id={c.type_id}, "
                        f"code={c.code}.",
            ))
    if existing_errors:
        return 0, existing_errors

    if dry_run:
        # No persistimos; preview OK
        return len(componentes), []

    # 3. Persistir transaccional
    try:
        with transaction.atomic():
            for c in componentes:
                # Resolver FKs a objetos para asignarlos
                district = District.objects.get(ubigeo=c.district_ubigeo)
                component_type = ComponentType.objects.get(pk=c.type_id)
                op_status = None
                fis_status = None
                if c.operational_status_code:
                    from components.models import OperationalStatus
                    op_status = OperationalStatus.objects.filter(
                        code=c.operational_status_code
                    ).first()
                if c.physical_status_code:
                    from components.models import PhysicalStatus
                    fis_status = PhysicalStatus.objects.filter(
                        code=c.physical_status_code
                    ).first()

                comp = Component.objects.create(
                    code=c.code,
                    name=c.name,
                    type=component_type,
                    district=district,
                    operational_status=op_status,
                    physical_status=fis_status,
                    specification=c.specification,
                )

                # Coords
                coord_instances = []
                for coord in c.coords:
                    if "easting" in coord and "northing" in coord:
                        point = SpatialHelper.utm_to_wgs84(
                            coord["easting"],
                            coord["northing"],
                            coord.get("srid_origin", 18),
                        )
                    else:
                        point = Point(coord["longitude"], coord["latitude"], srid=4326)
                    coord_instances.append(ComponentCoord(
                        component=comp,
                        criticality_id=coord["criticality"],
                        coords=point,
                    ))
                if coord_instances:
                    ComponentCoord.objects.bulk_create(coord_instances)
    except Exception as e:
        # Si es ValidationError de DRF, `e` puede venir con ErrorDetail
        # (objeto cuyo __repr__ es `ErrorDetail(string='...', code='...')`).
        # Lo aplanamos a un mensaje legible para el usuario final.
        msg: str
        if isinstance(e, ValidationError):
            # ValidationError.detail puede ser list / dict / str.
            detail = getattr(e, 'detail', None)
            if isinstance(detail, list):
                msg = "; ".join(str(item) for item in detail)
            elif isinstance(detail, dict):
                parts = []
                for k, v in detail.items():
                    if isinstance(v, list):
                        parts.append(f"{k}: " + "; ".join(str(x) for x in v))
                    else:
                        parts.append(f"{k}: {v}")
                msg = "; ".join(parts) if parts else str(e)
            else:
                msg = str(detail) if detail else str(e)
        else:
            msg = str(e)
        return 0, [ImportError(row=None, code=None, message=f"Error al persistir: {msg}")]

    return len(componentes), []